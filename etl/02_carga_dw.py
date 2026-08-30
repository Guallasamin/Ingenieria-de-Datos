#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=======================================================================
 PASO 2 - ETL DE CARGA DEL DATA WAREHOUSE (modelo estrella)
 Taller Individual Semana 1 | Ingenieria de Datos
=======================================================================
 Implementa exactamente la misma logica que el flujo de KNIME, para que
 el DW quede poblado y verificable aunque KNIME no este disponible.
 Cada bloque indica entre corchetes el nodo KNIME equivalente.

   EXTRACCION   4 fuentes heterogeneas
   TRANSFORMACION  limpieza + conformado + enriquecimiento + SCD2
   CARGA        6 dimensiones + 1 tabla de hechos

 Ejecucion:  docker exec etl_runtime python etl/02_carga_dw.py
=======================================================================
"""
import os, re, glob, unicodedata, datetime as dt
import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text

RAIZ     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR_DATA = os.path.join(RAIZ, "data")
CSV_2026 = os.path.join(RAIZ, "fuentes", "csv", "colocaciones_2026.csv")

PG = "postgresql+psycopg2://{u}:{p}@{h}:{P}/{d}".format(
    u=os.getenv("PG_USER","etl_user"), p=os.getenv("PG_PASS","etl_pass_2026"),
    h=os.getenv("PG_HOST","postgres"), P=os.getenv("PG_PORT","5432"),
    d=os.getenv("PG_DB","banca_ec"))
MY = "mysql+pymysql://{u}:{p}@{h}:{P}/{d}?charset=utf8mb4".format(
    u=os.getenv("MY_USER","etl_user"), p=os.getenv("MY_PASS","etl_pass_2026"),
    h=os.getenv("MY_HOST","mysql"),   P=os.getenv("MY_PORT","3306"),
    d=os.getenv("MY_DB","catalogos_sb"))

SEGMENTOS = {
    "BASE B PRIVADA CONSUMO":"CONSUMO", "BASE B PRIVADA EDUCATIVO":"EDUCATIVO",
    "BASE B PRIVADA MICROCREDITO":"MICROCREDITO", "BASE B PRIVADA PRODUCTIVO":"PRODUCTIVO",
    "BASE B PRIVADA INMOBILIARIO":"INMOBILIARIO",
    "BASE B PRIVADA VIVIENDA INTERES":"VIVIENDA INTERES PUBLICO"}

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
         "Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

def log(m): print(f"[{dt.datetime.now():%H:%M:%S}] {m}", flush=True)

# ---------------- funciones de limpieza (identicas al paso 1) ----------
def limpiar(s):
    return None if s is None else re.sub(r"\s+", " ", str(s).strip())
def sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")
def nombre_comercial(bruto):
    n = limpiar(bruto).upper()
    n = re.sub(r",?\s*EN\s+LIQUIDACION\s*$", "", n)
    for _ in range(2):
        n = re.sub(r"^(BP|BANCO)\s+", "", n)
    n = re.sub(r"\bS\.\s?A\.?(?=\s|,|$)", "", n)
    n = re.sub(r"\s+,", ",", limpiar(n))
    return limpiar(n).strip(" ,")
def estado_de(bruto):
    return "EN LIQUIDACION" if "LIQUIDACION" in limpiar(bruto).upper() else "ACTIVA"
def codigo(t, largo=30):
    c = re.sub(r"[^A-Z0-9]+", "_", sin_tildes(limpiar(t).upper())).strip("_")
    return c[:largo].rstrip("_")

def leer_hoja_base(ruta, hoja):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja]; cab=None; idx=None; filas=[]
    for fila in ws.iter_rows(values_only=True):
        v = list(fila)
        if cab is None:
            if any(x is not None and str(x).strip().upper()=="FECHA" for x in v):
                cab=[None if x is None else str(x).strip().upper() for x in v]
                idx={h:j for j,h in enumerate(cab) if h}
            continue
        if all(x is None for x in v): continue
        filas.append({h: v[j] for h,j in idx.items()})
    wb.close()
    return pd.DataFrame(filas)

# =====================================================================
def main():
    pg, my = create_engine(PG), create_engine(MY)
    log("="*68); log("PASO 2 - CARGA DEL DATA WAREHOUSE"); log("="*68)

    # =================================================================
    # E X T R A C C I O N
    # =================================================================
    log("EXTRACCION de las 4 fuentes")

    # --- FUENTE 1: PostgreSQL  [nodo: PostgreSQL Connector + DB Reader]
    capt = pd.read_sql(text("""
        SELECT fecha_corte, entidad, region, provincia, canton,
               cuenta_contable, tipo_deposito,
               numero_cuentas, numero_clientes, saldo
        FROM staging.captaciones_banca_privada"""), pg)
    log(f"   F1 PostgreSQL | captaciones .......... {len(capt):>7,} filas")

    # --- FUENTE 2: MySQL       [nodo: MySQL Connector + DB Reader x3]
    cat_ent  = pd.read_sql("SELECT * FROM cat_entidad_financiera", my)
    cat_geo  = pd.read_sql("SELECT * FROM cat_geografia", my)
    cat_prod = pd.read_sql("SELECT * FROM cat_producto_financiero", my)
    cat_ope  = pd.read_sql("SELECT * FROM cat_tipo_operacion", my)
    cat_fue  = pd.read_sql("SELECT * FROM cat_fuente_datos", my)
    log(f"   F2 MySQL      | catalogos ............ {len(cat_ent)}+{len(cat_geo)}"
        f"+{len(cat_prod)}+{len(cat_ope)}+{len(cat_fue)} filas")

    # --- FUENTE 3: Excel XLSX  [nodo: Excel Reader]
    partes = []
    for ruta in sorted(glob.glob(os.path.join(DIR_DATA, "202[45]", "Cartera", "*.xlsx"))):
        wb = openpyxl.load_workbook(ruta, read_only=True)
        hojas = [h for h in wb.sheetnames if h.upper().startswith("BASE")]; wb.close()
        for hoja in hojas:
            d = leer_hoja_base(ruta, hoja)
            d["SEGMENTO"] = SEGMENTOS.get(hoja.upper().strip(), hoja)
            partes.append(d)
    col_xlsx = pd.concat(partes, ignore_index=True)
    log(f"   F3 Excel XLSX | colocaciones 24-25 ... {len(col_xlsx):>7,} filas")

    # --- FUENTE 4: CSV         [nodo: CSV Reader]
    col_csv = pd.read_csv(CSV_2026, sep=";", encoding="utf-8")
    col_csv = col_csv.rename(columns={
        "fecha_corte":"FECHA", "entidad":"ENTIDAD", "provincia":"PROVINCIA",
        "canton":"CANTON", "segmento":"SEGMENTO", "por_vencer":"POR VENCER",
        "no_devenga_intereses":"NO DEVENGA INTERESES", "vencida":"VENCIDA",
        "total_saldo":"TOTAL SALDO"})
    log(f"   F4 CSV        | colocaciones 2026 .... {len(col_csv):>7,} filas")

    # =================================================================
    # T R A N S F O R M A C I O N
    # =================================================================
    log("-"*68); log("TRANSFORMACION")

    # T1. Union de las dos fuentes de colocaciones  [nodo: Concatenate]
    cols = ["FECHA","ENTIDAD","PROVINCIA","CANTON","SEGMENTO",
            "POR VENCER","NO DEVENGA INTERESES","VENCIDA","TOTAL SALDO"]
    col_xlsx["_fuente"] = "F3_EXCEL"; col_csv["_fuente"] = "F4_CSV"
    colo = pd.concat([col_xlsx[cols+["_fuente"]], col_csv[cols+["_fuente"]]],
                     ignore_index=True)
    log(f"   T1 Concatenate   Excel + CSV -> {len(colo):,} colocaciones")

    # T2. Limpieza de cadenas  [nodo: String Manipulation]
    #     colapsa dobles espacios, recorta, unifica mayusculas
    for d in (colo, capt):
        d["FECHA"] = pd.to_datetime(d["FECHA"] if "FECHA" in d else d["fecha_corte"])
        for c in [x for x in ("ENTIDAD","PROVINCIA","CANTON","entidad","provincia","canton") if x in d]:
            d[c] = d[c].map(limpiar)
    capt["tipo_deposito"] = capt.tipo_deposito.map(limpiar)
    n_dobles = int(col_xlsx.ENTIDAD.astype(str).str.contains(r"\s{2,}").sum() +
                   capt.entidad.astype(str).str.contains(r"\s{2,}").sum())
    log(f"   T2 Limpieza      espacios multiples normalizados")

    # T3. Conformado de la clave de entidad  [nodo: String Manipulation + Rule Engine]
    for d, c in ((colo,"ENTIDAD"), (capt,"entidad")):
        d["nombre_comercial"] = d[c].map(nombre_comercial)
        d["cod_entidad"]      = d.nombre_comercial.map(codigo)
        d["estado_entidad"]   = d[c].map(estado_de)
    log(f"   T3 Conformado    {colo.cod_entidad.nunique()} entidades en colocaciones / "
        f"{capt.cod_entidad.nunique()} en captaciones (claves unificadas)")

    # T4. Clave geografica  [nodo: String Manipulation]
    for d, p, c in ((colo,"PROVINCIA","CANTON"), (capt,"provincia","canton")):
        d["cod_geografia"] = (d[p].map(codigo) + "_" + d[c].map(codigo)).str.slice(0,60)

    # T5. Enriquecimiento: la REGION solo existe en la fuente PostgreSQL,
    #     se propaga a las colocaciones  [nodo: Joiner (left outer)]
    mapa_region = cat_geo.set_index("cod_geografia").region
    colo["region"] = colo.cod_geografia.map(mapa_region).fillna("NO DEFINIDA")
    sin_region = int((colo.region == "NO DEFINIDA").sum())
    log(f"   T5 Enriquecim.   region propagada a colocaciones "
        f"({sin_region} filas sin correspondencia -> 'NO DEFINIDA')")
    if sin_region:
        faltan = (colo.loc[colo.region == "NO DEFINIDA", ["PROVINCIA","CANTON"]]
                      .drop_duplicates().values.tolist())
        log(f"      cantones sin region en el catalogo: {faltan}")

    # T6. Clave de producto  [nodo: Rule Engine]
    colo["cod_producto"] = "CRED_" + colo.SEGMENTO.map(lambda s: codigo(s,30))
    capt["cod_producto"] = "DEP_"  + capt.tipo_deposito.map(lambda s: codigo(s,30))

    # T7. Metricas derivadas  [nodo: Math Formula]
    colo["saldo_improductivo"] = (colo["NO DEVENGA INTERESES"].astype(float)
                                  + colo["VENCIDA"].astype(float)).round(2)
    capt["saldo_promedio_cliente"] = (capt.saldo.astype(float) /
        capt.numero_clientes.astype(float).replace(0, float("nan"))).round(2)

    # T8. Control de calidad  [nodo: Rule-based Row Filter]
    #     descarta filas sin fecha, sin entidad o con saldo negativo
    antes_c, antes_d = len(colo), len(capt)
    colo = colo[colo.FECHA.notna() & colo.cod_entidad.ne("") &
                (colo["TOTAL SALDO"].astype(float) >= 0)]
    capt = capt[capt.FECHA.notna() & capt.cod_entidad.ne("") &
                (capt.saldo.astype(float) >= 0)]
    log(f"   T8 Calidad       rechazadas {antes_c-len(colo)} colocaciones / "
        f"{antes_d-len(capt)} captaciones")

    # =================================================================
    # C A R G A   D E   D I M E N S I O N E S   [nodo: DB Writer]
    # =================================================================
    log("-"*68); log("CARGA DE DIMENSIONES")
    with pg.begin() as cx:
        cx.execute(text("TRUNCATE TABLE dw.fact_saldos_financieros RESTART IDENTITY CASCADE"))
        for t_ in ("dim_tiempo","dim_entidad","dim_geografia","dim_producto",
                   "dim_tipo_operacion","dim_fuente_datos"):
            cx.execute(text(f"TRUNCATE TABLE dw.{t_} RESTART IDENTITY CASCADE"))

    # --- D1 dim_tiempo  [nodo: Date&Time-based Row Generator]
    fechas = sorted(set(colo.FECHA.dt.date) | set(capt.FECHA.dt.date))
    dim_t = pd.DataFrame({"fecha": fechas})
    f = pd.to_datetime(dim_t.fecha)
    dim_t["tiempo_sk"]       = f.dt.strftime("%Y%m%d").astype(int)
    dim_t["anio"]            = f.dt.year
    dim_t["semestre"]        = ((f.dt.month - 1)//6 + 1)
    dim_t["trimestre"]       = f.dt.quarter
    dim_t["mes"]             = f.dt.month
    dim_t["nombre_mes"]      = f.dt.month.map(lambda m: MESES[m-1])
    dim_t["nombre_mes_corto"]= dim_t.nombre_mes.str[:3]
    dim_t["anio_mes"]        = f.dt.strftime("%Y-%m")
    dim_t["etiqueta_trim"]   = f.dt.year.astype(str) + "-T" + f.dt.quarter.astype(str)
    dim_t["dia_del_mes"]     = f.dt.day
    dim_t["es_fin_trimestre"]= f.dt.month.isin([3,6,9,12])
    dim_t["es_fin_anio"]     = f.dt.month == 12
    dim_t["es_cierre_fiscal"]= f.dt.month == 12
    dim_t.to_sql("dim_tiempo", pg, schema="dw", if_exists="append", index=False)
    log(f"   D1 dim_tiempo ............ {len(dim_t):>4} cortes mensuales "
        f"({dim_t.anio_mes.min()} a {dim_t.anio_mes.max()})")

    # --- D2 dim_entidad con SCD TIPO 2  [nodo: DB Reader MySQL -> DB Writer]
    # Las versiones las define el catalogo maestro de MySQL (fuente 2), de modo
    # que el flujo de KNIME y este ETL producen exactamente la misma dimension.
    dim_e = cat_ent[["cod_entidad", "nombre_entidad", "nombre_comercial",
                     "estado_entidad", "grupo_tamanio", "perfil_negocio",
                     "cobertura_geo", "num_provincias", "num_cantones",
                     "version", "fecha_inicio_vig", "fecha_fin_vig",
                     "es_vigente"]].copy()
    dim_e["es_vigente"] = dim_e.es_vigente.astype(bool)
    dim_e.to_sql("dim_entidad", pg, schema="dw", if_exists="append", index=False)
    n_scd = int((dim_e.version > 1).sum())
    log(f"   D2 dim_entidad ........... {len(dim_e):>4} filas "
        f"({dim_e.cod_entidad.nunique()} entidades, {n_scd} version(es) historica(s) SCD2)")

    # --- D3 dim_geografia  [nodo: DB Reader MySQL -> DB Writer]
    dim_g = cat_geo.drop(columns=["geografia_sk"])
    for b in ("es_canton_principal","es_registro_desconocido"):
        dim_g[b] = dim_g[b].astype(bool)
    dim_g.to_sql("dim_geografia", pg, schema="dw", if_exists="append", index=False)
    log(f"   D3 dim_geografia ......... {len(dim_g):>4} cantones "
        f"({dim_g.region.nunique()} regiones, {dim_g.provincia.nunique()} provincias)")

    # --- D4 dim_producto
    dim_p = cat_prod.drop(columns=["producto_sk"])
    dim_p["es_a_la_vista"] = dim_p.es_a_la_vista.astype(bool)
    dim_p.to_sql("dim_producto", pg, schema="dw", if_exists="append", index=False)
    log(f"   D4 dim_producto .......... {len(dim_p):>4} productos")

    # --- D5 dim_tipo_operacion / D6 dim_fuente_datos
    cat_ope.drop(columns=["tipo_operacion_sk"]).to_sql("dim_tipo_operacion", pg, schema="dw", if_exists="append", index=False)
    cat_fue.drop(columns=["fuente_sk"]).to_sql("dim_fuente_datos",   pg, schema="dw", if_exists="append", index=False)
    log(f"   D5 dim_tipo_operacion .... {len(cat_ope):>4} filas")
    log(f"   D6 dim_fuente_datos ...... {len(cat_fue):>4} filas")

    # =================================================================
    # C A R G A   D E   H E C H O S   [nodo: Joiner x6 + DB Writer]
    # Sustitucion de claves de negocio por claves subrogadas
    # =================================================================
    log("-"*68); log("CARGA DE LA TABLA DE HECHOS")
    sk_t = pd.read_sql("SELECT tiempo_sk, fecha FROM dw.dim_tiempo", pg)
    sk_t["fecha"] = pd.to_datetime(sk_t.fecha)
    sk_e = pd.read_sql("""SELECT entidad_sk, cod_entidad, estado_entidad,
                                 fecha_inicio_vig, fecha_fin_vig FROM dw.dim_entidad""", pg)
    sk_e["fecha_inicio_vig"] = pd.to_datetime(sk_e.fecha_inicio_vig)
    sk_e["fecha_fin_vig"]    = pd.to_datetime(sk_e.fecha_fin_vig, errors="coerce").fillna(
                                pd.Timestamp("2262-01-01"))
    sk_g = pd.read_sql("SELECT geografia_sk, cod_geografia FROM dw.dim_geografia", pg)
    sk_p = pd.read_sql("SELECT producto_sk, cod_producto FROM dw.dim_producto", pg)
    sk_o = pd.read_sql("SELECT tipo_operacion_sk, cod_operacion FROM dw.dim_tipo_operacion", pg)
    sk_f = pd.read_sql("SELECT fuente_sk, cod_fuente FROM dw.dim_fuente_datos", pg)

    def resolver_entidad(df):
        """Lookup SCD2: elige la version vigente en la fecha del hecho."""
        m = df.merge(sk_e, on="cod_entidad", how="left")
        m = m[(m.FECHA >= m.fecha_inicio_vig) & (m.FECHA <= m.fecha_fin_vig)]
        return m.drop(columns=["fecha_inicio_vig","fecha_fin_vig","estado_entidad_y"],
                      errors="ignore")

    # ---- hechos de COLOCACIONES
    hc = colo.rename(columns={"POR VENCER":"saldo_por_vencer",
                              "NO DEVENGA INTERESES":"saldo_no_devenga",
                              "VENCIDA":"saldo_vencido",
                              "TOTAL SALDO":"saldo_total"})
    hc = resolver_entidad(hc)
    hc = (hc.merge(sk_t, left_on="FECHA", right_on="fecha", how="left")
            .merge(sk_g, on="cod_geografia", how="left")
            .merge(sk_p, on="cod_producto", how="left"))
    hc["tipo_operacion_sk"] = sk_o.loc[sk_o.cod_operacion=="COL","tipo_operacion_sk"].iloc[0]
    hc["fuente_sk"] = hc._fuente.map(sk_f.set_index("cod_fuente").fuente_sk)
    hc["numero_cuentas"] = pd.NA; hc["numero_clientes"] = pd.NA
    hc["saldo_promedio_cliente"] = pd.NA

    # ---- hechos de CAPTACIONES
    hd = capt.rename(columns={"saldo":"saldo_total"})
    hd = resolver_entidad(hd)
    hd = (hd.merge(sk_t, left_on="FECHA", right_on="fecha", how="left")
            .merge(sk_g, on="cod_geografia", how="left")
            .merge(sk_p, on="cod_producto", how="left"))
    hd["tipo_operacion_sk"] = sk_o.loc[sk_o.cod_operacion=="CAP","tipo_operacion_sk"].iloc[0]
    hd["fuente_sk"] = sk_f.loc[sk_f.cod_fuente=="F1_POSTGRES","fuente_sk"].iloc[0]
    for c in ("saldo_por_vencer","saldo_no_devenga","saldo_vencido","saldo_improductivo"):
        hd[c] = pd.NA

    COLS = ["tiempo_sk","entidad_sk","geografia_sk","producto_sk","tipo_operacion_sk",
            "fuente_sk","saldo_total","saldo_por_vencer","saldo_no_devenga",
            "saldo_vencido","saldo_improductivo","numero_cuentas","numero_clientes",
            "saldo_promedio_cliente"]
    hechos = pd.concat([hc[COLS], hd[COLS]], ignore_index=True)

    # Integridad referencial: ninguna FK puede quedar nula  [nodo: Missing Value + Row Filter]
    huerfanos = hechos[hechos[["tiempo_sk","entidad_sk","geografia_sk",
                               "producto_sk"]].isna().any(axis=1)]
    if len(huerfanos):
        log(f"   !! {len(huerfanos)} filas huerfanas descartadas (FK sin correspondencia)")
        hechos = hechos.drop(huerfanos.index)
    for c in ("tiempo_sk","entidad_sk","geografia_sk","producto_sk",
              "tipo_operacion_sk","fuente_sk"):
        hechos[c] = hechos[c].astype(int)

    hechos.to_sql("fact_saldos_financieros", pg, schema="dw", if_exists="append",
                  index=False, chunksize=10000, method="multi")
    log(f"   FACT fact_saldos_financieros ... {len(hechos):,} filas cargadas")

    with pg.begin() as cx:
        cx.execute(text("""INSERT INTO staging.etl_bitacora
          (proceso,fuente,objeto_destino,filas_leidas,filas_escritas,filas_rechazadas,
           estado,mensaje,fin)
          VALUES ('02_carga_dw','4 FUENTES','dw.fact_saldos_financieros',
                  :l,:e,:r,'OK','Carga completa del modelo estrella',now())"""),
          {"l": int(antes_c+antes_d), "e": int(len(hechos)),
           "r": int(antes_c+antes_d-len(hechos))})

    log("="*68); log("PASO 2 COMPLETADO - Data Warehouse poblado."); log("="*68)

if __name__ == "__main__":
    main()
