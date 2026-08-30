#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=======================================================================
 PASO 1 - PREPARACION DE LAS 4 FUENTES HETEROGENEAS
 Taller Individual Semana 1 | Ingenieria de Datos
=======================================================================
 Lee los libros Excel originales de la Superintendencia de Bancos y
 distribuye la informacion en cuatro tecnologias distintas, para que el
 flujo ETL de KNIME deba integrarlas realmente:

   FUENTE 1  PostgreSQL   staging.captaciones_banca_privada (122.217 filas)
   FUENTE 2  MySQL        catalogos_sb.* (5 catalogos maestros derivados)
   FUENTE 3  Excel XLSX   data/2024|2025/Cartera/*.xlsx  (se usan tal cual)
   FUENTE 4  CSV          fuentes/csv/colocaciones_2026.csv

 Ejecucion:  docker exec etl_runtime python etl/01_extraccion_fuentes.py
=======================================================================
"""
import os, re, sys, glob, unicodedata, datetime as dt
from collections import defaultdict

import pandas as pd
import openpyxl
from sqlalchemy import create_engine, text

# ---------------------------------------------------------------- rutas
RAIZ     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIR_DATA = os.path.join(RAIZ, "data")
DIR_CSV  = os.path.join(RAIZ, "fuentes", "csv")
os.makedirs(DIR_CSV, exist_ok=True)

# ------------------------------------------------------- conexiones BD
PG = "postgresql+psycopg2://{u}:{p}@{h}:{P}/{d}".format(
    u=os.getenv("PG_USER", "etl_user"), p=os.getenv("PG_PASS", "etl_pass_2026"),
    h=os.getenv("PG_HOST", "postgres"), P=os.getenv("PG_PORT", "5432"),
    d=os.getenv("PG_DB",   "banca_ec"))
MY = "mysql+pymysql://{u}:{p}@{h}:{P}/{d}?charset=utf8mb4".format(
    u=os.getenv("MY_USER", "etl_user"), p=os.getenv("MY_PASS", "etl_pass_2026"),
    h=os.getenv("MY_HOST", "mysql"),   P=os.getenv("MY_PORT", "3306"),
    d=os.getenv("MY_DB",   "catalogos_sb"))

# Nombre de hoja BASE -> segmento de credito normalizado
SEGMENTOS = {
    "BASE B PRIVADA CONSUMO":         "CONSUMO",
    "BASE B PRIVADA EDUCATIVO":       "EDUCATIVO",
    "BASE B PRIVADA MICROCREDITO":    "MICROCREDITO",
    "BASE B PRIVADA PRODUCTIVO":      "PRODUCTIVO",
    "BASE B PRIVADA INMOBILIARIO":    "INMOBILIARIO",
    "BASE B PRIVADA VIVIENDA INTERES":"VIVIENDA INTERES PUBLICO",
}

def log(msg):
    print(f"[{dt.datetime.now():%H:%M:%S}] {msg}", flush=True)

# =====================================================================
#  UTILIDADES DE LIMPIEZA  (se replican como nodos String Manipulation
#  en el flujo KNIME - ver knime/GUIA_FLUJO_KNIME.md)
# =====================================================================
def limpiar(s):
    """Recorta y colapsa espacios multiples. Resuelve el caso real
       'BP BANCO  DESARROLLO DE LOS PUEBLOS  S.A., CODESARROLLO'."""
    if s is None:
        return None
    return re.sub(r"\s+", " ", str(s).strip())

def sin_tildes(s):
    """Normalizacion NFD: quita diacriticos para construir claves estables."""
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")

def nombre_comercial(bruto):
    """Nombre comercial homogeneo y comparable entre las 4 fuentes.

    Aplica, en orden: colapso de espacios -> separacion del estado juridico
    -> eliminacion del prefijo del regulador ('BP', 'BANCO', que aparecen
    incluso duplicados) -> eliminacion de la forma societaria ('S.A.').
    Ej.: 'BP BANCO  DESARROLLO DE LOS PUEBLOS  S.A., CODESARROLLO'
         -> 'DESARROLLO DE LOS PUEBLOS, CODESARROLLO'
    """
    n = limpiar(bruto).upper()
    n = re.sub(r",?\s*EN\s+LIQUIDACION\s*$", "", n)   # el estado no es parte del nombre
    for _ in range(2):                                  # prefijos (pueden venir duplicados)
        n = re.sub(r"^(BP|BANCO)\s+", "", n)
    n = re.sub(r"\bS\.\s?A\.?(?=\s|,|$)", "", n)      # forma societaria
    n = re.sub(r"\s+,", ",", limpiar(n))                # comas huerfanas
    return limpiar(n).strip(" ,")

def estado_de(bruto):
    return "EN LIQUIDACION" if "LIQUIDACION" in limpiar(bruto).upper() else "ACTIVA"

def codigo(texto, largo=30):
    """Clave de negocio: mayusculas, sin tildes, sin simbolos."""
    c = sin_tildes(limpiar(texto).upper())
    c = re.sub(r"[^A-Z0-9]+", "_", c).strip("_")
    return c[:largo].rstrip("_")

# =====================================================================
#  LECTOR GENERICO DE LAS HOJAS "BASE ..." DE LA SUPERINTENDENCIA
#  Las hojas traen filas y columnas vacias de relleno; localizamos la
#  cabecera buscando la celda literal 'FECHA'.
# =====================================================================
def leer_hoja_base(ruta, hoja):
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[hoja]
    cabecera, idx, filas = None, None, []
    for fila in ws.iter_rows(values_only=True):
        vals = list(fila)
        if cabecera is None:
            if any(v is not None and str(v).strip().upper() == "FECHA" for v in vals):
                cabecera = [None if v is None else str(v).strip().upper() for v in vals]
                idx = {h: j for j, h in enumerate(cabecera) if h}
            continue
        if all(v is None for v in vals):
            continue
        filas.append({h: vals[j] for h, j in idx.items()})
    wb.close()
    return pd.DataFrame(filas)

def leer_cartera(anios):
    """Consolida las hojas BASE de todos los libros de Cartera."""
    partes = []
    for ruta in sorted(glob.glob(os.path.join(DIR_DATA, "*", "Cartera", "*.xlsx"))):
        anio = int(os.path.basename(os.path.dirname(os.path.dirname(ruta))))
        if anio not in anios:
            continue
        wb = openpyxl.load_workbook(ruta, read_only=True)
        hojas = [h for h in wb.sheetnames if h.upper().startswith("BASE")]
        wb.close()
        for hoja in hojas:
            df = leer_hoja_base(ruta, hoja)
            if df.empty:
                continue
            df["SEGMENTO"]       = SEGMENTOS.get(hoja.upper().strip(), hoja)
            df["ARCHIVO_ORIGEN"] = os.path.basename(ruta)
            df["ANIO_ARCHIVO"]   = anio
            partes.append(df)
            log(f"   Cartera {anio} | {hoja:<34} {len(df):>6,} filas")
    return pd.concat(partes, ignore_index=True)

def leer_depositos():
    partes = []
    for ruta in sorted(glob.glob(os.path.join(DIR_DATA, "*", "Depositos", "*.xlsx"))):
        anio = int(os.path.basename(os.path.dirname(os.path.dirname(ruta))))
        df = leer_hoja_base(ruta, "BASE BANCA PRIVADA")
        df["ARCHIVO_ORIGEN"] = os.path.basename(ruta)
        df["ANIO_ARCHIVO"]   = anio
        partes.append(df)
        log(f"   Depositos {anio} | {len(df):>7,} filas")
    return pd.concat(partes, ignore_index=True)

# =====================================================================
#  MAIN
# =====================================================================
def main():
    log("=" * 68)
    log("PASO 1 - PREPARACION DE LAS 4 FUENTES")
    log("=" * 68)

    # ---------------------------------------------------- lectura Excel
    log("Leyendo libros de CARTERA (colocaciones)...")
    cartera = leer_cartera({2024, 2025, 2026})
    log("Leyendo libros de DEPOSITOS (captaciones)...")
    depositos = leer_depositos()
    log(f"TOTAL leido -> cartera {len(cartera):,} | depositos {len(depositos):,}")

    # ------------------------------------- tipificacion y normalizacion
    for df in (cartera, depositos):
        df["FECHA"] = pd.to_datetime(df["FECHA"]).dt.date
        for col in ("ENTIDAD", "PROVINCIA", "CANTON"):
            df[col] = df[col].map(limpiar)
    depositos["REGION"]        = depositos["REGION"].map(limpiar)
    depositos["TIPO DE DEPOSITO"] = depositos["TIPO DE DEPOSITO"].map(limpiar)
    depositos["CUENTA"]        = depositos["CUENTA"].astype(str).str.strip()

    # =================================================================
    #  FUENTE 1 -> PostgreSQL  (staging.captaciones_banca_privada)
    # =================================================================
    log("-" * 68)
    log("FUENTE 1 | Cargando CAPTACIONES en PostgreSQL...")
    pg = create_engine(PG)
    dep_pg = depositos.rename(columns={
        "FECHA": "fecha_corte", "ENTIDAD": "entidad", "REGION": "region",
        "PROVINCIA": "provincia", "CANTON": "canton", "CUENTA": "cuenta_contable",
        "TIPO DE DEPOSITO": "tipo_deposito", "NUMERO DE CUENTAS": "numero_cuentas",
        "NUMERO DE CLIENTES": "numero_clientes", "SALDO": "saldo",
        "ARCHIVO_ORIGEN": "archivo_origen", "ANIO_ARCHIVO": "anio_archivo"})[[
        "fecha_corte", "entidad", "region", "provincia", "canton", "cuenta_contable",
        "tipo_deposito", "numero_cuentas", "numero_clientes", "saldo",
        "archivo_origen", "anio_archivo"]]
    with pg.begin() as cx:
        cx.execute(text("TRUNCATE TABLE staging.captaciones_banca_privada RESTART IDENTITY"))
    dep_pg.to_sql("captaciones_banca_privada", pg, schema="staging",
                  if_exists="append", index=False, chunksize=10000, method="multi")
    n = pd.read_sql("SELECT COUNT(*) c FROM staging.captaciones_banca_privada", pg).c[0]
    log(f"   -> staging.captaciones_banca_privada: {n:,} filas")

    # =================================================================
    #  FUENTE 4 -> CSV  (colocaciones 2026)
    # =================================================================
    log("-" * 68)
    log("FUENTE 4 | Exportando COLOCACIONES 2026 a CSV...")
    c26 = cartera[cartera.ANIO_ARCHIVO == 2026].rename(columns={
        "FECHA": "fecha_corte", "ENTIDAD": "entidad", "PROVINCIA": "provincia",
        "CANTON": "canton", "POR VENCER": "por_vencer",
        "NO DEVENGA INTERESES": "no_devenga_intereses", "VENCIDA": "vencida",
        "TOTAL SALDO": "total_saldo", "SEGMENTO": "segmento",
        "ARCHIVO_ORIGEN": "archivo_origen"})[[
        "fecha_corte", "entidad", "provincia", "canton", "segmento",
        "por_vencer", "no_devenga_intereses", "vencida", "total_saldo",
        "archivo_origen"]]
    destino_csv = os.path.join(DIR_CSV, "colocaciones_2026.csv")
    c26.to_csv(destino_csv, index=False, sep=";", encoding="utf-8", decimal=".")
    log(f"   -> {destino_csv}: {len(c26):,} filas")

    # =================================================================
    #  FUENTE 2 -> MySQL  (catalogos maestros DERIVADOS del perfilado)
    # =================================================================
    log("-" * 68)
    log("FUENTE 2 | Derivando y cargando CATALOGOS en MySQL...")
    my = create_engine(MY)
    # idempotencia: el script debe poder re-ejecutarse sin duplicar catalogos
    with my.begin() as cx:
        for t_ in ("cat_entidad_financiera", "cat_geografia", "cat_producto_financiero"):
            cx.execute(text(f"DELETE FROM {t_}"))

    # ---- catalogo de entidades -------------------------------------
    ent = pd.DataFrame({"nombre_entidad": sorted(
        set(cartera.ENTIDAD) | set(depositos.ENTIDAD))})
    ent["nombre_comercial"] = ent.nombre_entidad.map(nombre_comercial)
    ent["estado_entidad"]   = ent.nombre_entidad.map(estado_de)
    ent["cod_entidad"]      = ent.nombre_comercial.map(codigo)

    # tamanio: participacion acumulada en captaciones (criterio de Pareto)
    vol = (depositos.assign(e=depositos.ENTIDAD.map(nombre_comercial))
                    .groupby("e").SALDO.sum().sort_values(ascending=False))
    acum = vol.cumsum() / vol.sum()
    def tamanio(nc):
        a = acum.get(nc)
        if a is None:   return "PEQUENO"
        if a <= 0.70:   return "GRANDE"
        if a <= 0.90:   return "MEDIANO"
        return "PEQUENO"
    ent["grupo_tamanio"] = ent.nombre_comercial.map(tamanio)

    # perfil de negocio: segmento dominante de la cartera de cada entidad
    mix = (cartera.assign(e=cartera.ENTIDAD.map(nombre_comercial))
                  .groupby(["e", "SEGMENTO"])["TOTAL SALDO"].sum().unstack(fill_value=0))
    share = mix.div(mix.sum(axis=1).replace(0, pd.NA), axis=0)
    def perfil(nc):
        if nc not in share.index: return "MIXTO"
        f = share.loc[nc]
        viv = f.get("INMOBILIARIO", 0) + f.get("VIVIENDA INTERES PUBLICO", 0)
        cand = {"MICROFINANZAS": f.get("MICROCREDITO", 0), "CONSUMO": f.get("CONSUMO", 0),
                "COMERCIAL": f.get("PRODUCTIVO", 0), "VIVIENDA": viv}
        k = max(cand, key=cand.get)
        return k if cand[k] >= 0.50 else "MIXTO"
    ent["perfil_negocio"] = ent.nombre_comercial.map(perfil)

    # cobertura territorial
    geo_ent = (pd.concat([cartera[["ENTIDAD", "PROVINCIA", "CANTON"]],
                          depositos[["ENTIDAD", "PROVINCIA", "CANTON"]]])
                 .assign(e=lambda d: d.ENTIDAD.map(nombre_comercial)))
    nprov = geo_ent.groupby("e").PROVINCIA.nunique()
    ncant = geo_ent.groupby("e").CANTON.nunique()
    ent["num_provincias"] = ent.nombre_comercial.map(nprov).fillna(0).astype(int)
    ent["num_cantones"]   = ent.nombre_comercial.map(ncant).fillna(0).astype(int)
    ent["cobertura_geo"]  = pd.cut(ent.num_provincias, [-1, 4, 14, 99],
                                   labels=["LOCAL", "REGIONAL", "NACIONAL"]).astype(str)

    # fecha de alta = primer corte en que la entidad aparece reportando
    primera = (pd.concat([cartera[["ENTIDAD", "FECHA"]], depositos[["ENTIDAD", "FECHA"]]])
                 .groupby("ENTIDAD").FECHA.min())
    ent["fecha_alta"] = ent.nombre_entidad.map(primera)

    # --- versiones SCD Tipo 2: una fila por (entidad, estado observado) ----
    obs = pd.concat([
        cartera.assign(nc=cartera.ENTIDAD.map(nombre_comercial),
                       est=cartera.ENTIDAD.map(estado_de))[["ENTIDAD","nc","est","FECHA"]],
        depositos.assign(nc=depositos.ENTIDAD.map(nombre_comercial),
                         est=depositos.ENTIDAD.map(estado_de))[["ENTIDAD","nc","est","FECHA"]]])
    vig = (obs.groupby(["nc", "est", "ENTIDAD"]).FECHA.min()
              .reset_index(name="fecha_inicio_vig"))
    vig["cod_entidad"] = vig.nc.map(codigo)
    vig = vig.sort_values(["cod_entidad", "fecha_inicio_vig"])
    vig["version"] = vig.groupby("cod_entidad").cumcount() + 1
    sig = vig.groupby("cod_entidad").fecha_inicio_vig.shift(-1)
    vig["fecha_fin_vig"] = [
        (x - dt.timedelta(days=1)) if pd.notna(x) else dt.date(9999, 12, 31)
        for x in sig]
    vig["es_vigente"] = sig.isna().astype(int)
    cat_ent = (vig.rename(columns={"nc": "nombre_comercial", "est": "estado_entidad",
                                   "ENTIDAD": "nombre_entidad"})
                  .merge(ent[["cod_entidad", "grupo_tamanio", "perfil_negocio",
                              "cobertura_geo", "num_provincias", "num_cantones"]]
                         .drop_duplicates("cod_entidad"),
                         on="cod_entidad", how="left"))
    cat_ent["fecha_alta"] = cat_ent.fecha_inicio_vig
    cat_ent = cat_ent.sort_values(["cod_entidad", "version"]).reset_index(drop=True)
    cat_ent.insert(0, "entidad_sk", range(1, len(cat_ent) + 1))
    cat_ent = cat_ent[["entidad_sk", "cod_entidad", "version", "fecha_inicio_vig",
                       "fecha_fin_vig", "es_vigente", "nombre_entidad",
                       "nombre_comercial", "estado_entidad", "grupo_tamanio",
                       "perfil_negocio", "cobertura_geo", "num_provincias",
                       "num_cantones", "fecha_alta"]]
    cat_ent.to_sql("cat_entidad_financiera", my, if_exists="append", index=False)
    log(f"   -> cat_entidad_financiera: {len(cat_ent)} filas "
        f"({cat_ent.cod_entidad.nunique()} entidades, "
        f"{int((cat_ent.version > 1).sum())} version(es) SCD2)")

    # ---- catalogo geografico ---------------------------------------
    geo = (pd.concat([
              cartera[["PROVINCIA", "CANTON", "ENTIDAD"]].assign(REGION=pd.NA),
              depositos[["PROVINCIA", "CANTON", "ENTIDAD", "REGION"]]]))
    # la REGION solo existe en Depositos -> se propaga a Cartera por provincia
    mapa_region = (depositos.dropna(subset=["REGION"])
                            .groupby("PROVINCIA").REGION.agg(lambda s: s.mode()[0]))
    cat_geo = (geo.groupby(["PROVINCIA", "CANTON"])
                  .ENTIDAD.nunique().reset_index(name="num_entidades_presentes"))
    cat_geo["region"] = cat_geo.PROVINCIA.map(mapa_region).fillna("NO DEFINIDA")
    cat_geo["cod_geografia"] = (cat_geo.PROVINCIA.map(codigo) + "_" +
                                cat_geo.CANTON.map(codigo)).map(lambda s: s[:60])
    cat_geo["nivel_bancarizacion"] = pd.cut(
        cat_geo.num_entidades_presentes, [-1, 3, 9, 99],
        labels=["BAJA", "MEDIA", "ALTA"]).astype(str)
    cat_geo["es_registro_desconocido"] = (cat_geo.region == "NO DEFINIDA").astype(int)
    # canton principal = el de mayor saldo total dentro de su provincia
    vol_cant = (pd.concat([
        cartera.groupby(["PROVINCIA", "CANTON"])["TOTAL SALDO"].sum(),
        depositos.groupby(["PROVINCIA", "CANTON"]).SALDO.sum()]).groupby(level=[0, 1]).sum())
    principal = vol_cant.groupby(level=0).idxmax()
    setp = set(principal.values)
    cat_geo["es_canton_principal"] = [
        int((p, c) in setp) for p, c in zip(cat_geo.PROVINCIA, cat_geo.CANTON)]
    cat_geo = (cat_geo.rename(columns={"PROVINCIA": "provincia", "CANTON": "canton"})
                      .sort_values("cod_geografia").reset_index(drop=True))
    cat_geo.insert(0, "geografia_sk", range(1, len(cat_geo) + 1))
    cat_geo = cat_geo[[
        "geografia_sk", "cod_geografia", "canton", "provincia", "region",
        "nivel_bancarizacion", "num_entidades_presentes", "es_canton_principal",
        "es_registro_desconocido"]]
    cat_geo.to_sql("cat_geografia", my, if_exists="append", index=False)
    log(f"   -> cat_geografia: {len(cat_geo)} cantones "
        f"({cat_geo.es_registro_desconocido.sum()} sin region en origen)")

    # ---- catalogo de productos -------------------------------------
    prods = []
    for i, seg in enumerate(sorted(cartera.SEGMENTO.unique()), start=1):
        prods.append(dict(cod_producto="CRED_" + codigo(seg, 30), nombre_producto=seg,
                          familia="CARTERA", subfamilia="CREDITO " + seg,
                          cuenta_contable=None, es_a_la_vista=0,
                          plazo_dias_min=None, plazo_dias_max=None, orden_presentacion=i))
    PLAZOS = {"De 1 a 30 dias": (1, 30), "De 31 a 90 dias": (31, 90),
              "De 91 a 180 dias": (91, 180), "De 181 A 360 dias": (181, 360),
              "De mas De 361 dias": (361, None)}
    tip = (depositos[["TIPO DE DEPOSITO", "CUENTA"]].drop_duplicates()
                    .groupby("TIPO DE DEPOSITO").CUENTA.min())
    for j, (nom, cta) in enumerate(tip.items(), start=100):
        clave = sin_tildes(nom)
        pmin, pmax = PLAZOS.get(clave, (None, None))
        es_plazo = clave in PLAZOS
        prods.append(dict(cod_producto="DEP_" + codigo(nom, 30), nombre_producto=nom,
                          familia="DEPOSITO",
                          subfamilia="PLAZO FIJO" if es_plazo else "A LA VISTA",
                          cuenta_contable=str(cta), es_a_la_vista=0 if es_plazo else 1,
                          plazo_dias_min=pmin, plazo_dias_max=pmax, orden_presentacion=j))
    cat_prod = pd.DataFrame(prods).sort_values("orden_presentacion").reset_index(drop=True)
    cat_prod.insert(0, "producto_sk", range(1, len(cat_prod) + 1))
    cat_prod.to_sql("cat_producto_financiero", my, if_exists="append", index=False)
    log(f"   -> cat_producto_financiero: {len(cat_prod)} productos "
        f"({(cat_prod.familia=='CARTERA').sum()} credito / "
        f"{(cat_prod.familia=='DEPOSITO').sum()} deposito)")

    # =================================================================
    #  FUENTE 3 -> Excel (verificacion, los .xlsx se leen tal cual)
    # =================================================================
    log("-" * 68)
    x = sorted(glob.glob(os.path.join(DIR_DATA, "202[45]", "Cartera", "*.xlsx")))
    log(f"FUENTE 3 | Excel XLSX disponible: {len(x)} libros 2024-2025 "
        f"({len(cartera[cartera.ANIO_ARCHIVO.isin([2024,2025])]):,} filas)")

    # ---- bitacora ---------------------------------------------------
    with pg.begin() as cx:
        cx.execute(text("""INSERT INTO staging.etl_bitacora
            (proceso,fuente,objeto_destino,filas_leidas,filas_escritas,estado,mensaje,fin)
            VALUES ('01_extraccion','F1_POSTGRES','staging.captaciones_banca_privada',
                    :n,:n,'OK','Carga de captaciones desde libros Excel',now())"""),
            {"n": int(len(dep_pg))})
        cx.execute(text("""INSERT INTO staging.etl_bitacora
            (proceso,fuente,objeto_destino,filas_leidas,filas_escritas,estado,mensaje,fin)
            VALUES ('01_extraccion','F4_CSV','fuentes/csv/colocaciones_2026.csv',
                    :n,:n,'OK','Exportacion de colocaciones 2026',now())"""),
            {"n": int(len(c26))})
        cx.execute(text("""INSERT INTO staging.etl_bitacora
            (proceso,fuente,objeto_destino,filas_leidas,filas_escritas,estado,mensaje,fin)
            VALUES ('01_extraccion','F2_MYSQL','catalogos_sb.*',
                    :n,:n,'OK','Catalogos maestros derivados del perfilado',now())"""),
            {"n": int(len(cat_ent) + len(cat_geo) + len(cat_prod))})

    log("=" * 68)
    log("PASO 1 COMPLETADO - las 4 fuentes estan disponibles.")
    log("=" * 68)

if __name__ == "__main__":
    main()
